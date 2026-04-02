"""Recalculate Excel formulas using LibreOffice.

Installs a LibreOffice macro on first run, then executes it to recalculate
all formulas and scan for errors.

Usage:
    python recalc.py <excel_file> [timeout_seconds]

Returns JSON with:
    - status: 'success' or 'errors_found'
    - total_errors: count of Excel errors
    - total_formulas: number of formulas in file
    - error_summary: breakdown by error type with cell locations
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

try:
    from office.soffice import get_soffice_env
except ImportError:
    def get_soffice_env():
        env = os.environ.copy()
        env["SAL_USE_VCLPLUGIN"] = "svp"
        return env

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def setup_macro() -> bool:
    macro_dir = os.path.expanduser(
        MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX
    )
    macro_file = os.path.join(macro_dir, "Module1.xba")

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    if not os.path.exists(macro_dir):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True, timeout=10, check=False, env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename: str, timeout: int = 30) -> dict:
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin":
        try:
            subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1, check=False)
            cmd = ["gtimeout", str(timeout)] + cmd
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    if result.returncode not in (0, 124):
        return {"error": result.stderr or "Unknown error during recalculation"}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"status": "recalculated", "note": "Install openpyxl to scan for errors"}

    # Scan for errors
    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_details = {e: [] for e in excel_errors}
    total_errors = 0

    wb = load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for err in excel_errors:
                        if err in cell.value:
                            error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb.close()

    # Count formulas
    wb2 = load_workbook(filename, data_only=False)
    formula_count = sum(
        1 for sn in wb2.sheetnames for row in wb2[sn].iter_rows()
        for cell in row
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("=")
    )
    wb2.close()

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
    }

    if total_errors > 0:
        result["error_summary"] = {
            err: {"count": len(locs), "locations": locs[:20]}
            for err, locs in error_details.items() if locs
        }

    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("Recalculates all formulas and scans for errors. Returns JSON.")
        sys.exit(1)

    fname = sys.argv[1]
    tout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalc(fname, tout), indent=2))
